import streamlit as st
import io
from datetime import datetime
import pandas as pd
import streamlit_authenticator as stauth
from time import sleep
import json 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dependencies import (
    consulta_nombre,
    consulta_general,
    add_registro,
    crear_tabla,
    crear_tabla_proyectos,
    add_registro_proyecto,
    consulta_proyectos,
    consulta_por_codigo,
    add_material_usado,
    get_materiales_por_proyecto,
    resumen_materiales,
    obtener_proyecto_id,
    obtener_metas_hh, 
    guardar_meta_hh,
    delete_materiales_por_proyecto_fecha)
st.set_page_config(page_title="Libro Diario de Obra", layout="centered",page_icon="logo_epesa2.png")
st.markdown("""
    <style>
    /* Fondo general del sidebar */
    section[data-testid="stSidebar"] {
        background-color: #0055A4;
    }

    /* ✅ Solo estos elementos en blanco */
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] h4,
    section[data-testid="stSidebar"] h5,
    section[data-testid="stSidebar"] h6,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stText,
    section[data-testid="stSidebar"] .stMarkdown {
        color: white !important;
    }

    /* Selectbox fondo blanco */
    section[data-testid="stSidebar"] div[role="combobox"] {
        background-color: white !important;
        border-radius: 6px;
        padding: 0.25rem 0.5rem;
    }

    /* ✅ Valor actual en negro */
    section[data-testid="stSidebar"] div[role="combobox"] > div:first-child span {
        color: black !important;
        font-weight: 500 !important;
    }

    /* Ícono flecha del selectbox */
    section[data-testid="stSidebar"] div[role="combobox"] svg {
        fill: black !important;
    }

    /* Opciones desplegadas */
    ul[role="listbox"] li {
        color: black !important;
    }
    </style>
""", unsafe_allow_html=True)


# ---- Personalización de botones ----
st.markdown(
    """
    <style>
    /* Estilo base */
    .stButton > button {
        background-color: #003366;   /* Azul oscuro */
        color: #FFFFFF;              /* Texto blanco */
        border-radius: 0.5rem;       /* Esquinas redondeadas */
        padding: 0.75rem 1.5rem;     /* Altura y anchura del botón */
        font-size: 1rem;             /* Tamaño de fuente */
        font-weight: 600;            /* Negrita ligera */
        transition: background-color 0.2s ease;
    }
    /* Hover */
    .stButton > button:hover {
        background-color: #002244;
    }
    /* Active/Click */
    .stButton > button:active {
        background-color: #001122;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ——————————————————————————————————————————
# 0) Pre‑carga global del catálogo con cache y filtrado
# ——————————————————————————————————————————
@st.cache_data(show_spinner=False)
def load_catalogo(path: str) -> dict:
    import pandas as pd

    # Leer columnas necesarias usando nombres actualizados
    try:
        df = pd.read_excel(
            path,
            sheet_name='BASE',
            usecols=[
                "Número de artículo", 
                "Descripción del artículo", 
                "Grupo de artículos", 
                "Unidad medida compras",
                "Costo del artículo"
            ]
        )
    except Exception as e:
        st.error(f"❌ Error al leer el catálogo: {e}")
        return {}

    # Filtrar artículos que no nos interesan
    df["Grupo de artículos"] = df["Grupo de artículos"].astype(str)
    df = df[~df["Grupo de artículos"].str.upper().str.startswith("HERRAMIENTAS")]

    # Limpieza básica
    df = df.dropna(subset=["Número de artículo", "Descripción del artículo"])
    df["Número de artículo"] = df["Número de artículo"].astype(str)

    # Normalizar campos faltantes
    df["Unidad medida compras"] = df["Unidad medida compras"].fillna("UN")
    df["Costo del artículo"] = df["Costo del artículo"].fillna(0.0)

    # Etiqueta de tipo general
    df["tipo"] = df["Grupo de artículos"].str.upper().str.strip()

    # Convertir a diccionario por código
    catalogo = {
        row["Número de artículo"]: {
            "descripcion": row["Descripción del artículo"],
            "unidad": row["Unidad medida compras"],
            "tipo": row["tipo"]
            
        }
        for _, row in df.iterrows()
    }

    return catalogo

catalogo_dict = load_catalogo("LISTADO DE ARTICULOS SAP.xlsx")
 
# ——————————————————————————————————————————

COOKIE_EXPIRY_DAYS = 30  # Días de expiración de la cookie,MANTIENE AL USUARIO LOGUEADO
# Autenticación de usuario
def main():

    try:
        consulta_general()
    except Exception:
        crear_tabla()

    db_query = consulta_general()

    registros={'usernames':{}}
    for data in db_query:
        registros['usernames'][data[1]] = {
            'name': data[0],
            'password': data[2]
        }

    authenticator = stauth.Authenticate(
        registros,
        'random_cookie_name',
        'random_signature_key',
        COOKIE_EXPIRY_DAYS,
    )
    if 'click_en_registro' not in st.session_state:
        st.session_state['click_en_registro'] = False
    if st.session_state['click_en_registro']==False:
        login_form(authenticator=authenticator)
    else:
        usuario_form()

# Garantiza existencia sin duplicar nada
crear_tabla_proyectos()
import streamlit as st
from datetime import datetime

def proceso(authenticator, name, username):
    authenticator.logout("Logout", 'main')
    st.title(f"Bienvenido {name}!")

    # Lista de actividades y roles
    actividades = [
    "Ducteado Embutido/Endosado",
    "Bandejado",
    "Cableado",
    "Mecanismos",
    "Artefactos de Iluminación",
    "MT",
    "Excavación",
    "Tableros Eléctricos",
    "Puesta a Tierra",
    "Pararrayos"   
]

    roles = ["Oficial", "Ayudante", "Contratista"]

    # 1. Encabezado
    with st.expander("1. Encabezado", expanded=True):
        c1, c2, c3 = st.columns(3)
        fecha    = c1.date_input("Fecha", datetime.today(), key="fecha")
        cliente  = c2.text_input("Cliente", key="cliente")
        proyecto = c3.text_input("Proyecto (Código)", key="proyecto")

    # 2. Condiciones Climáticas
    with st.expander("2. Condiciones Climáticas", expanded=False):
        m1, m2 = st.columns(2)
        clima_manana = m1.radio("Mañana", ["Soleado", "Nublado", "Lluvioso"], key="clima_manana", horizontal=True)
        clima_tarde  = m2.radio("Tarde",   ["Soleado", "Nublado", "Lluvioso"], key="clima_tarde",  horizontal=True)


    def reset_activity(act_prefix):
        """
        Al marcar el checkbox de reset para la actividad:
        - Buscamos todas las claves de session_state que empiecen
        con act_prefix + '_'
        - Si terminan en '_cnt' o '_hh' las ponemos a 0
        si terminan en '_obs' las ponemos a ''
        - Finalmente, desmarcamos el propio checkbox
        """
        to_clear = []
        for k in st.session_state.keys():
            if k.startswith(act_prefix + "_"):
                to_clear.append(k)
        for k in to_clear:
            if k.endswith("_cnt"):
                st.session_state[k] = 0
            elif k.endswith("_hh"):
                st.session_state[k] = 0.0
            elif k.endswith("_obs"):
                st.session_state[k] = ""
        # desmarcar el checkbox
        st.session_state[f"{act_prefix}_reset_all"] = False


    st.subheader("Personal Directo")
    registro_actividades = []

    for act in actividades:
        with st.expander(act, expanded=False):
            # 1) Checkbox de reset con callback
            reset_key = f"{act}_reset_all"
            if reset_key not in st.session_state:
                st.session_state[reset_key] = False

            st.checkbox(
                "Restablecer todos los valores de esta actividad",
                key=reset_key,
                help="Marca para limpiar todos los campos de esta actividad",
                on_change=reset_activity,
                args=(act,)
            )

            # 2) Ahora construye tus tabs (o zonas) sabiendo que,
            #    si el checkbox se marcó, en el callback ya limpiamos todo
            if act == "Ducteado Embutido/Endosado":
                zonas = ["MAMPOSTERIA", "LOSA", "SUBTERRANEO", "EXTERNO"]
                zona_tabs = st.tabs(zonas)
                for zona, tab in zip(zonas, zona_tabs):
                    with tab:
                        datos = {"Actividad": f"{zona}-{act}", "Roles": {}, "Obs": ""}
                        roles = ["Oficial", "Ayudante", "Contratista"]
                        role_tabs = st.tabs(roles)
                        for role, rt in zip(roles, role_tabs):
                            with rt:
                                col1, col2 = st.columns(2)
                                cnt_key = f"{act}_{zona}_{role}_cnt"
                                hh_key  = f"{act}_{zona}_{role}_hh"
                                # valores por defecto vienen del session_state (o 0 tras reset)
                                cnt = col1.number_input(
                                    "Personal",
                                    min_value=0,
                                    key=cnt_key
                                )
                                hh = col2.number_input(
                                    "Horas trabajadas por persona",
                                    min_value=0.0, step=0.5,
                                    key=hh_key
                                )
                                datos["Roles"][role] = {"Personal": cnt, "HH": hh}

                        obs_key = f"{act}_{zona}_obs"
                        obs = st.text_input(
                            "Observaciones generales",
                            key=obs_key
                        )
                        datos["Obs"] = obs
                        registro_actividades.append(datos)
            # Cableado con dos subdivisiones
            elif act == "Cableado":
                zonas = ["CABLEADO CIRCUITOS", "ALIMENTADORES CABLES"]
                zona_tabs = st.tabs(zonas)
                for zona, tab in zip(zonas, zona_tabs):
                    with tab:
                        datos = {"Actividad": f"{zona}-{act}", "Roles": {}, "Obs": ""}
                        roles = ["Oficial", "Ayudante", "Contratista"]
                        role_tabs = st.tabs(roles)
                        for role, rt in zip(roles, role_tabs):
                            with rt:
                                col1, col2 = st.columns(2)
                                cnt_key = f"{act}_{zona}_{role}_cnt"
                                hh_key  = f"{act}_{zona}_{role}_hh"
                                cnt = col1.number_input("Personal", min_value=0, key=cnt_key)
                                hh  = col2.number_input("Horas trabajadas por persona", min_value=0.0, step=0.5, key=hh_key)
                                datos["Roles"][role] = {"Personal": cnt, "HH": hh}

                        obs_key = f"{act}_{zona}_obs"
                        datos["Obs"] = st.text_input("Observaciones generales", key=obs_key)
                        registro_actividades.append(datos)
            elif act == "Tableros Eléctricos":
                # Tableros con zonas
                zonas = ["Fabricación Tableros", "Conexión de Tableros"]
                zona_tabs = st.tabs(zonas)
                for zona, tab in zip(zonas, zona_tabs):
                    with tab:
                        datos = {"Actividad": f"{zona}-{act}", "Roles": {}, "Obs": ""}
                        roles = ["Oficial Tablerista", "Ayudante Tablerista", "Contratista"]
                        role_tabs = st.tabs(roles)
                        for role, rt in zip(roles, role_tabs):
                            with rt:
                                col1, col2 = st.columns(2)
                                cnt_key = f"{act}_{zona}_{role}_cnt"
                                hh_key  = f"{act}_{zona}_{role}_hh"
                                cnt = col1.number_input("Personal", min_value=0, key=cnt_key)
                                hh  = col2.number_input("Horas trabajadas por persona", min_value=0.0, step=0.5, key=hh_key)
                                datos["Roles"][role] = {"Personal": cnt, "HH": hh}

                        obs_key = f"{act}_{zona}_obs"
                        datos["Obs"] = st.text_input("Observaciones generales", key=obs_key)
                        registro_actividades.append(datos)
            elif act == "Bandejado":
                zonas = ["Soporteria", "Bandejado"]
                zona_tabs = st.tabs(zonas)
                for zona, tab in zip(zonas, zona_tabs):
                    with tab:
                        datos = {"Actividad": f"{zona}-{act}", "Roles": {}, "Obs": ""}
                        roles = ["Oficial", "Ayudante", "Contratista"]
                        role_tabs = st.tabs(roles)
                        for role, rt in zip(roles, role_tabs):
                            with rt:
                                col1, col2 = st.columns(2)
                                cnt_key = f"{act}_{zona}_{role}_cnt"
                                hh_key  = f"{act}_{zona}_{role}_hh"
                                cnt = col1.number_input("Personal", min_value=0, key=cnt_key)
                                hh  = col2.number_input("Horas trabajadas por persona", min_value=0.0, step=0.5, key=hh_key)
                                datos["Roles"][role] = {"Personal": cnt, "HH": hh}

                        # campo de observaciones (igual que en Tableros)
                        obs_key = f"{act}_{zona}_obs"
                        datos["Obs"] = st.text_input("Observaciones generales", key=obs_key)

                        registro_actividades.append(datos)

            else:
                # actividades normales (sin zonas)
                roles_act = (
                    ["Oficial Tablerista", "Ayudante Tablerista", "Contratista"]
                    if act == "Tableros Eléctricos"
                    else ["Oficial", "Ayudante", "Contratista"]
                )
                role_tabs = st.tabs(roles_act)
                datos = {"Actividad": act, "Roles": {}, "Obs": ""}
                for role, rt in zip(roles_act, role_tabs):
                    with rt:
                        col1, col2 = st.columns(2)
                        cnt_key = f"{act}_{role}_cnt"
                        hh_key  = f"{act}_{role}_hh"
                        cnt = col1.number_input(
                            "Personal",
                            min_value=0,
                            key=cnt_key
                        )
                        hh = col2.number_input(
                            "Horas trabajadas por persona",
                            min_value=0.0, step=0.5,
                            key=hh_key
                        )
                        datos["Roles"][role] = {"Personal": cnt, "HH": hh}
                obs_key = f"{act}_obs"
                datos["Obs"] = st.text_input("Observaciones generales", key=obs_key)
                registro_actividades.append(datos)

# Ahora `registro_actividades` viene sin necesidad de borrar manualmente cada campo,
# y el checkbox de reset siempre queda listo para volver a usarse.

    # 4. Detalle
    detalle = st.text_area("Detalle de las actividades realizadas", height=120, key="detalle")

    
    # 5. Equipos en obra (dinámico)
    # 5. Equipos en obra (dinámico con borrado por fila)
    st.subheader("Equipos en Obra")

    # 1) Inicializar lista de equipos en sesión (si no existe)
    if "equipos" not in st.session_state:
        # empezamos con 3 filas vacías
        st.session_state.equipos = [{"tipo": "", "cantidad": 0} for _ in range(3)]

    # 2) Botón para añadir más equipos
    if st.button("➕ Añadir más equipos"):
        st.session_state.equipos.append({"tipo": "", "cantidad": 0})

    # 3) Para cada equipo en sesión, mostramos sus inputs + botón de borrar
    equipos = []
    for idx, equipo in enumerate(st.session_state.equipos):
        col1, col2, col3 = st.columns([3, 1, 0.5])
        tipo = col1.text_input(
            f"Tipo de equipo {idx+1}",
            value=equipo["tipo"],
            key=f"eq_tipo_{idx}"
        )
        cantidad = col2.number_input(
            f"Cantidad {idx+1}",
            min_value=0,
            value=equipo["cantidad"],
            key=f"eq_cant_{idx}"
        )
        # botón de borrar fila
        if col3.button("🗑️", key=f"eq_del_{idx}"):
            # quitamos este índice y refrescamos
            st.session_state.equipos.pop(idx)
            st.rerun()
        else:
            # si no borramos, lo guardamos en la nueva lista
            equipos.append({"tipo": tipo, "cantidad": cantidad})

    # 4) Actualizar la sesión con las filas no borradas
    st.session_state.equipos = equipos

    # 6. Firmas
    residente  = st.text_input("Firma Residente - Nombre", key="firma_residente")
    encargado  = st.text_input("Firma Encargado - Nombre", key="firma_encargado")

    # BOTÓN DE ENVÍO (¡¡solo esto dispara el registro, no el Enter!!)
    # 0) Construir el dict de registro antes de cualquier botón
    registro = {
        "fecha":        fecha,
        "cliente":      cliente,
        "proyecto":     proyecto,
        "clima_manana": clima_manana,
        "clima_tarde":  clima_tarde,
        "detalle":      detalle,
        "actividades":  registro_actividades,
        "equipos":      equipos,
        "residente":    residente,
        "encargado":    encargado,
        "usuario":      username,
    }

   # 1) Control de estado para mostrar el resumen
    if "show_review" not in st.session_state:
        st.session_state.show_review = False

    # 2) Botón inicial para doble check
    if not st.session_state.show_review:
        if st.button("Confirmar registro", key="btn_review"):
            st.session_state.show_review = True

    # 3) Si el usuario activó la revisión, mostramos el doble check y el confirm
    if st.session_state.show_review:
        st.markdown("### ⚠️ Revisa lo que cargaste antes de confirmar")
        st.markdown(f"- **Usuario que registra:** {registro['usuario']}")
        st.markdown(f"- **Fecha:** {registro['fecha'].strftime('%d/%m/%Y')}")
        st.markdown(f"- **Cliente:** {registro['cliente']}")
        st.markdown(f"- **Proyecto:** {registro['proyecto']}")
        st.markdown(f"- **Clima:** {registro['clima_manana']} / {registro['clima_tarde']}")
        st.markdown("#### 📝 Detalle narrativo")
        st.write(registro["detalle"])
        st.markdown("#### 🖋️Firmas")
        st.markdown(f"- **Residente:** {registro['residente']}")
        st.markdown(f"- **Encargado:** {registro['encargado']}")
        st.markdown("#### 👷 Actividades realizadas")
        for act in registro["actividades"]:
            total_pers = sum(v["Personal"] for v in act["Roles"].values())
            total_hh   = sum(v["Personal"] * v["HH"] for v in act["Roles"].values())
            if total_pers == 0 and total_hh == 0:
                continue
            st.markdown(f"**• {act['Actividad']}**")
            df_roles = pd.DataFrame.from_dict(act["Roles"], orient="index") \
                            .rename(columns={"Personal":"Cant. Pers.","HH":"HH p/p"})
            st.table(df_roles)
            if act.get("Obs"):
                st.markdown(f"_Observaciones:_ {act['Obs']}")

        st.markdown("#### 🚜 Equipos en obra")

        equipos_raw = registro.get("equipos", [])

        if equipos_raw:
            df_equipos = pd.DataFrame(equipos_raw)

            if "tipo" in df_equipos.columns:
                # Filtrar filas con valor válido en 'tipo'
                df_equipos = df_equipos[df_equipos["tipo"].str.strip() != ""]

                if not df_equipos.empty:
                    df_equipos = df_equipos.rename(columns={
                        "tipo": "Tipo de equipo",
                        "cantidad": "Cantidad"
                    })

                    df_equipos.index = range(1, len(df_equipos) + 1)
                    st.table(df_equipos)
                else:
                    st.info("🚫 No hay equipos cargados.")
            else:
                st.info("🚫 No hay equipos cargados.")
        else:
            st.info("🚫 No hay equipos cargados.")



        # 4) Botón definitivo
        if st.button("✅ Confirmar registro", key="btn_confirm"):
            # 4.1) Validar firmas
            if not registro["residente"].strip() or not registro["encargado"].strip():
                st.error("❌ No se puede guardar el registro. Completá los campos de **Residente** y **Encargado**.")
            else:
                # 4.2) Preparamos la fecha en ISO para comparar
                fecha_iso = registro["fecha"].isoformat()

                # 4.3) Obtenemos todos los registros existentes de este proyecto
                registros_existentes = consulta_por_codigo(registro["proyecto"])

                # 4.4) Comprobamos si ya existe uno con la misma fecha
                ya_existe = any(r[2] == fecha_iso for r in registros_existentes)

                # 4.5) Inicializar el flag de sobrescritura si no existe
                if "overwrite_confirmed" not in st.session_state:
                    st.session_state.overwrite_confirmed = False

                # 4.6) Si ya existe y aún no confirmaste overwrite, mostramos warning y paramos
                if ya_existe and not st.session_state.overwrite_confirmed:
                    st.warning(
                        "⚠️ Ya existe un registro para este proyecto y esta fecha.\n\n"
                        "Podés consultarlo desde 🔍 *Consultar registros*.\n\n"
                        "Si de todas formas querés sobrescribirlo, volvé a presionar **Confirmar registro**."
                    )
                    st.session_state.overwrite_confirmed = True
                    st.stop()

                # 4.7) Si no existe o ya confirmaste overwrite, grabamos
                with st.spinner("⏳ Registrando..."):
                    add_registro_proyecto(registro)

                st.success(
                    f"Registro del proyecto “{registro['proyecto']}” correspondiente al "
                    f"{registro['fecha'].strftime('%d/%m/%Y')} ha sido guardado con éxito."
                )

                # 4.8) Limpiamos estados para la próxima vez
                st.session_state.show_review = False
                st.session_state.overwrite_confirmed = False




        # Función para generar Excel
        def generar_excel(fecha, cliente, proyecto, clima_manana, clima_tarde, actividades, detalle, equipos):
            # Encabezado
            df_enc = pd.DataFrame({
                "Fecha": [fecha],
                "Cliente": [cliente],
                "Proyecto": [proyecto],
                "Clima Mañana": [clima_manana],
                "Clima Tarde": [clima_tarde],
                "Residente": [residente],
                "Encargado": [encargado]
            })
            # Actividades por rol
            rows = []

            import json

            # Asegurarse que `actividades` sea una lista de dicts
            if isinstance(actividades, str):
                try:
                    actividades = json.loads(actividades)
                except json.JSONDecodeError:
                    st.error("El campo 'actividades' no contiene un JSON válido")
                    actividades = []

            rows = []
            
            for act in actividades:
                if isinstance(act, str):
                    try:
                        act = json.loads(act)
                    except json.JSONDecodeError:
                        st.warning("Actividad con formato inválido (no es JSON válido)")
                        continue  # saltamos este act

                if not isinstance(act, dict) or "Roles" not in act:
                    st.warning("Formato inesperado en actividad")
                    continue

                for role, vals in act["Roles"].items():
                    rows.append({
                        "Actividad": act["Actividad"],
                        "Rol": role,
                        "Personal": vals["Personal"],
                        "HH": vals["HH"],
                        "Obs": act.get("Obs", "")
                    })

            df_act = pd.DataFrame(rows)
            df_act["HH Total"] = df_act["Personal"] * df_act["HH"]
            df_det = pd.DataFrame({"Detalle": [detalle]})
            df_eq = pd.DataFrame(equipos)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_enc.to_excel(writer, sheet_name="Encabezado", index=False)
                df_act.to_excel(writer, sheet_name="Actividades", index=False)
                df_det.to_excel(writer, sheet_name="Detalle", index=False)
                df_eq.to_excel(writer, sheet_name="Equipos", index=False)
            return output.getvalue()


        # Mostrar resultados, gráficos y descarga
        import plotly.graph_objects as go

        # Gráficos: totales por actividad
        #Suma d Personal por Actividad
        p_sum = [
            sum(v["Personal"] for v in act["Roles"].values())
            for act in registro_actividades
        ]
        labels = [act["Actividad"] for act in registro_actividades]

        # Cálculo de Horas-Hombre Totales (Personal × Horas por persona)
        hh_total = [
            sum(vals["Personal"] * vals["HH"] for vals in act["Roles"].values())
            for act in registro_actividades
        ]

        # ——————————————————————————————————————————
        # Gráfico de HH Totales y Personal, sin nulos
        # ——————————————————————————————————————————

        # Partimos de tus listas originales
        # labels: List[str], hh_total: List[float | None], p_sum: List[int | None]

        # 1) Filtramos solo las actividades con HH válidas (> 0 y no None)
        labels_hh = []
        hh_f = []
        for lbl, hh in zip(labels, hh_total):
            if hh is not None and hh > 0:
                labels_hh.append(lbl)
                hh_f.append(hh)

        # 2) Filtramos solo las actividades con Personal válidos (> 0 y no None)
        labels_p = []
        p_f = []
        for lbl, p in zip(labels, p_sum):
            if p is not None and p > 0:
                labels_p.append(lbl)
                p_f.append(p)

        # 3) Dibujamos el bar chart de HH Totales
        fig = go.Figure(data=[
            go.Bar(
                x=labels_hh,
                y=hh_f,
                marker_color='crimson',
                hovertemplate="%{x}: %{y} HH<extra></extra>"
            )
        ])
        fig.update_layout(
            title=f"Horas-Hombre Totales por Actividad del {fecha}",
            xaxis_title="Actividad",
            yaxis_title="HH Totales",
            xaxis_tickangle=-45,
            margin=dict(t=50, b=150)
        )
        st.plotly_chart(fig, use_container_width=True)

        # 4) Gráfico interactivo de Personal por Actividad
        fig2 = go.Figure(data=[
            go.Bar(
                x=labels_p,
                y=p_f,
                marker_color='teal',
                hovertemplate="%{x}: %{y} personas<extra></extra>"
            )
        ])
        fig2.update_layout(
            title=f"Total Personal por Actividad del {fecha}",
            xaxis_title="Actividad",
            yaxis_title="Personal",
            xaxis_tickangle=-45,
            margin=dict(t=50, b=150)
        )
        st.plotly_chart(fig2, use_container_width=True)


        # Descargar Excel
        excel_bytes = generar_excel(
            fecha, cliente, proyecto,
            clima_manana, clima_tarde,
            registro_actividades, detalle, equipos
        )
        st.download_button(
            label="📥 Descargar Excel de planilla diaria",
            data=excel_bytes,
            file_name=f"libro_obra_{proyecto}_{fecha}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        

def generar_excel_proyecto(codigo_id: str) -> bytes:
    """
    Genera un Excel con:
    - Hoja "Resumen" → tabla con fecha, cliente, clima, residente, encargado, usuario
    - Una hoja por fecha → detalle de actividades, equipos y observaciones
    - Hoja "Totales" → sumarios agregados de HH y personal por actividad
    """
    # 1) Traer todos los registros
    registros = consulta_por_codigo(codigo_id)
    
    # 2) Preparar datos de resumen y detalle por fecha
    resumen_rows = []
    detalle_por_fecha = {}
    actividades_totales = []
    
    for (
        codigo, cliente, fecha,
        clima_manana, clima_tarde,
        detalle, actividades, equipos,
        residente, encargado, usuario
    ) in registros:
        # Fila de resumen
        resumen_rows.append({
            "Fecha": fecha,
            "Cliente": cliente,
            "Clima Mañana": clima_manana,
            "Clima Tarde": clima_tarde,
            "Residente": residente,
            "Encargado": encargado,
            "Usuario": usuario
        })
        # Guardar info diaria
        detalle_por_fecha[fecha] = {
            "detalle": detalle,
            "actividades": actividades,
            "equipos": equipos
        }
        
        # Acumular para totales (convertimos actividades JSON)
        acts = actividades
        if isinstance(acts, str):
            try:
                acts = json.loads(acts)
            except json.JSONDecodeError:
                acts = []
        if not isinstance(acts, list):
            acts = []
        for act in acts:
            if isinstance(act, str):
                try:
                    act = json.loads(act)
                except:
                    continue
            if not isinstance(act, dict) or "Roles" not in act:
                continue
            nombre = act.get("Actividad", "")
            for role, vals in act["Roles"].items():
                actividades_totales.append({
                    "Actividad": nombre,
                    "Rol": role,
                    "Personal": vals.get("Personal", 0),
                    "HH": vals.get("HH", 0)
                })
    
    df_resumen = pd.DataFrame(resumen_rows).sort_values("Fecha")
    df_totales = (
        pd.DataFrame(actividades_totales)
          .groupby("Actividad", as_index=False)
          .agg({"Personal": "sum", "HH": "sum"})
    )
    
        

    # 3) Escribir todo en un solo ExcelWriter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hoja Resumen
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        
        # Hojas por cada fecha
        for fecha, info in detalle_por_fecha.items():
            sheet_name = fecha.isoformat() if hasattr(fecha, "isoformat") else str(fecha)
            
            # Construir df_act
            rows = []
            acts = info.get("actividades", []) or []
            if isinstance(acts, str):
                try: acts = json.loads(acts)
                except: acts = []
            for act in acts:
                if isinstance(act, str):
                    try: act = json.loads(act)
                    except: continue
                if not isinstance(act, dict) or "Roles" not in act:
                    continue
                obs    = act.get("Obs", "")
                nombre = act.get("Actividad", "Sin nombre")
                for role, vals in act["Roles"].items():
                    rows.append({
                        "Actividad": nombre,
                        "Rol":        role,
                        "Personal":   vals.get("Personal", 0),
                        "HH":         vals.get("HH",       0),
                        "Obs":        obs
                    })
            df_act = pd.DataFrame(rows)
            # Añadimos HH Totales en cada fila
            df_act["HH Totales"] = df_act["Personal"] * df_act["HH"]
            
            # Equipos
            eqs = info.get("equipos", []) or []
            if isinstance(eqs, str):
                try: eqs = json.loads(eqs)
                except: eqs = []
            df_eq = pd.DataFrame(eqs)
            
            # Volcar al sheet de la fecha
            df_act.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            df_eq.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(df_act) + 2)
            
            # Detalle narrativo
            ws = writer.book[sheet_name]
            ws.cell(row=len(df_act) + len(df_eq) + 5, column=1, value="Detalle:")
            ws.cell(row=len(df_act) + len(df_eq) + 6, column=1, value=info["detalle"])
        
        # Hoja Totales: agrupamos sobre todas las entradas
        df_all = pd.DataFrame(actividades_totales)
        # Asegurarnos de que tiene HH Totales
        df_all["HH Totales"] = df_all["Personal"] * df_all["HH"]
        # Agrupamos por Actividad, sumando Personal y HH Totales
        df_totales = (
            df_all
              .groupby("Actividad", as_index=False)
              .agg({
                  "Personal":   "sum",
                  "HH Totales": "sum"
              })
        )
        # Escribimos la hoja Totales
        df_totales.to_excel(writer, sheet_name="TotalesHH", index=False)
        
        # Reordenar hojas: poner "Totales" como segunda hoja
        sheets = writer.book.worksheets
        totales_sheet = sheets[-1]
        writer.book._sheets.remove(totales_sheet)
        writer.book._sheets.insert(1, totales_sheet)


        # ——————————————————————————————————————————————————
        # Hoja de MATERIALES
        # ——————————————————————————————————————————————————
        
        try:
            proyecto_id = obtener_proyecto_id(codigo_id)
            materiales = get_materiales_por_proyecto(proyecto_id) or []
            df_mat = pd.DataFrame(materiales)

            if not df_mat.empty:
                # renombrar columnas (ahora con Actividad)
                df_mat = df_mat.rename(columns={
                    "codigo": "ID",
                    "descripcion": "Nombre de tarea",
                    "unidad_medida": "U.M.",
                    "cantidad": "CANT.",
                    "tipo": "TIPO",
                    "actividad": "Actividad",
                    "fecha": "FECHA DE REGISTRO"
                })

                # seleccionar orden de columnas, incluyendo Actividad
                columnas = [
                    "ID", "Nombre de tarea", "U.M.", "CANT.",
                    "TIPO", "Actividad", "FECHA DE REGISTRO", "usuario"
                ]
                df_mat = df_mat[columnas]

                # vuelca el dataframe
                df_mat.to_excel(writer, sheet_name="Materiales", index=False)

                # aplicar estilos
                wb = writer.book
                ws = writer.sheets["Materiales"]

                header_fill = PatternFill("solid", fgColor="002060")
                header_font = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin = Border(
                    left=Side("thin"), right=Side("thin"),
                    top=Side("thin"), bottom=Side("thin")
                )

                # estilizar encabezado
                for col_idx, _ in enumerate(columnas, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.alignment = center
                    cell.border    = thin

                # bordes y formato numérico en datos
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columnas)):
                    for cell in row:
                        cell.border = thin
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

                # ajustar anchos
                for col_cells in ws.columns:
                    col_letter = get_column_letter(col_cells[0].column)
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
                    ws.column_dimensions[col_letter].width = max_len + 2

                # reposicionar la hoja
                mat_sheet = wb._sheets.pop(-1)
                wb._sheets.insert(2, mat_sheet)

        except Exception as e:
            print("Error al generar hoja de materiales:", e)

    output.seek(0)
    return output.getvalue()



from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def generar_planilla_excel(materiales, obra, cod_proj, fecha, buffer: io.BytesIO):
    wb = Workbook()
    ws = wb.active
    ws.title = "PLANILLA DE SEGUIMIENTO"

    # 🎨 Estilos
    header_fill = PatternFill("solid", fgColor="002060")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )

    # 1) Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "PLANILLA DE SEGUIMIENTO"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # 2) Metadatos
    ws["A2"] = "OBRA:";    ws["B2"] = obra or ""
    ws["A3"] = "COD:";     ws["B3"] = cod_proj
    ws["A4"] = "FECHA:";   ws["B4"] = str(fecha)

    # 3) Encabezados ajustados a tu tabla
    headers = ["Código", "Descripción", "Cantidad", "U.M.", "Tipo", "Actividad", "Subactividad"]
    ws.append([])  # fila en blanco
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # 4) Filas de materiales (sin agrupamientos)
    for m in materiales:
        ws.append([
            m.get("codigo", ""),
            m.get("descripcion", ""),
            m.get("cantidad", 0),
            m.get("unidad_medida", ""),
            m.get("tipo", ""),
            m.get("actividad", ""),
            m.get("subactividad", "")
        ])
        # aplicar bordes y formato numérico
        for cell in ws[ws.max_row]:
            cell.border = thin
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # 5) Ajustar anchos de columnas
    for col_cells in ws.columns:
        col = get_column_letter(col_cells[0].column)
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col].width = max_len + 2

    # 6) Guardar en el buffer
    wb.save(buffer)
    buffer.seek(0)




def login_form(authenticator):
    """
    Muestra el formulario de inicio de sesión y, tras autenticar,
    da acceso a registrar o consultar por código de proyecto.
    """
    name, authentication_status, username = authenticator.login("Login")

    if authentication_status:
        st.sidebar.success(f"Sesión: {name}", icon="✅")
        menu = st.sidebar.selectbox(
    "Menú",
    ["📝 Registrar nuevo", "🔍 Consultar registros",'🛠️ Registro de Material']
)

        if menu == "📝 Registrar nuevo":
            st.title("📝 LIBRO DIARIO DE OBRA")
            proceso(authenticator, name, username)  # tu lógica existente de guardado

        elif menu == "🛠️ Registro de Material":
            from datetime import date
            import io


            st.title("🛠️ REGISTRO DE MATERIAL")
            st.markdown("Registrá el material consumido en la obra.")

            # 1. Seleccionar proyecto
            codigos = consulta_proyectos()
            # Insertamos al inicio el texto guía
            codigos.insert(0, "Seleccione una obra")

            codigo_sel = st.selectbox(
                "Seleccionar obra / proyecto",
                codigos,
                index=0  # fuerza el placeholder como selección inicial
            )

            if codigo_sel == "Seleccione una obra":
                st.warning("Seleccioná un proyecto para continuar.")
                st.stop()

            # 2. Obtener ID interno
            try:
                proyecto_id = obtener_proyecto_id(codigo_sel)
            except Exception as e:
                st.error(f"No se pudo obtener el ID del proyecto: {e}")
                st.stop()


            # 3. Fecha de registro
            fecha_reg = st.date_input("📅 Fecha de registro", value=date.today())
            fecha_str = fecha_reg.isoformat()

            # 3.1 Mostrar materiales ya registrados en el proyecto 
            materiales_previos = get_materiales_por_proyecto(proyecto_id) or []
            df_prev = pd.DataFrame(materiales_previos)

            with st.expander("📦 Ver historial de materiales ya registrados en esta obra", expanded=False):
                if df_prev.empty:
                    st.info("No se cargaron materiales aún para este proyecto.")
                else:
                    # Renombrar columnas para visualización
                    df_prev = df_prev.rename(columns={
                        "codigo": "Código",
                        "descripcion": "Descripción",
                        "unidad_medida": "U.M.",
                        "cantidad": "Cantidad",
                        "tipo": "Tipo",
                        "fecha": "Fecha"
                    })
                    # Mostrar tabla
                    df_mostrar = df_prev[["Fecha", "Código", "Descripción", "Cantidad", "U.M.", "Tipo", "usuario"]] \
                        .sort_values("Fecha", ascending=False) \
                        .reset_index(drop=True)

                    df_mostrar.index += 1  # 👈 Esto cambia el índice a partir de 1

                    st.dataframe(df_mostrar, use_container_width=True)
                    st.caption("📤 Para descargar el historial completo, dirigite a la pestaña **📑 Consultar registros**.")

            # 4. Comprobamos si ya hay materiales para este proyecto+fecha
            existentes = get_materiales_por_proyecto(proyecto_id, fecha_str) or []
            if existentes and "confirm_overwrite_mat" not in st.session_state:
                st.warning(
                    "⚠️ Ya existe un registro de materiales para esta fecha.\n\n"
                    "Si querés reemplazarlo completamente, presioná **Guardar materiales** de nuevo."
                )
                st.session_state["confirm_overwrite_mat"] = False

            # 5. Inicializar lista de carga en sesión (vacía o recuperada)
            if "materiales_registro" not in st.session_state:
                st.session_state["materiales_registro"] = []

            # <<< NUEVO: Área para pegar códigos o nombres uno por línea
            st.markdown("#### Importar varios materiales de golpe")
            bulk_input = st.text_area(
                "Pega aquí CÓDIGOS o DESCRIPCIONES, uno por línea",
                placeholder="158000030\n158000032\nCANO CORRUGADO ANTILLAMA 25MM\" x 25MTS TIGREFLEX\n..."
            )
            if st.button("📥 Importar lista"):
                nuevos = []
                for linea in bulk_input.splitlines():
                    key = linea.strip()
                    if not key:
                        continue
                    # 1) ¿Es código válido?
                    if key in catalogo_dict:
                        cod = key
                    else:
                        # 2) Si no, buscar por descripción exacta (case-insensitive)
                        matches = [c for c,d in catalogo_dict.items()
                                if d["descripcion"].lower() == key.lower()]
                        cod = matches[0] if matches else None

                    if cod:
                        nuevos.append({
                            "codigo":       cod,
                            "descripcion":  catalogo_dict[cod]["descripcion"],
                            "cantidad":     0.0,
                            "unidad_medida":"UN",
                            "actividad":    "A elegir...",
                            "subactividad": ""
                        })
                if nuevos:
                    # Reemplaza la lista actual con los importados
                    st.session_state["materiales_registro"] = nuevos
                    st.rerun()
                else:
                    st.warning("No se encontraron coincidencias en el catálogo.")

            # 6. Botón para añadir fila nueva (dentro de un form)
            with st.form("add_material_form", clear_on_submit=True):
                add = st.form_submit_button("➕ Añadir otro material")

            if add:
                st.session_state["materiales_registro"].append({
                    "codigo": None,
                    "cantidad": 0.0,
                    "unidad_medida": "UN",
                    "actividad": "A elegir..."
                })
                st.rerun()


            # 7. Formulario dinámico con eliminación por fila
            materiales = st.session_state.get("materiales_registro", [])
            nuevos_materiales = []
            # Diccionario de subactividades
            subactividades_dict = {
                "Ducteado Embutido/Endosado": ["MAMPOSTERIA", "LOSA", "SUBTERRANEO", "EXTERNO"],
                "Cableado": ["CABLEADO CIRCUITOS", "ALIMENTADORES CABLES"],
                "Tableros Eléctricos": ["Fabricación Tableros", "Conexión de Tableros"],
                "Bandejado": ["Soporteria", "Bandejado"]
            }

            actividades = [
                "Ducteado Embutido/Endosado", "Bandejado", "Cableado", "Mecanismos",
                "Artefactos de Iluminación", "MT", "Excavación", "Tableros Eléctricos",
                "Puesta a Tierra", "Pararrayos"
            ]

            for idx, mat in enumerate(materiales):
                st.markdown(f"#### Material {idx+1}")
                col1, col2, col3, col4 = st.columns([3, 1.2, 1.2, 1])

                # Recuperar valores previos o usar por defecto
                codigo_prev = mat.get("codigo", None)
                cantidad_prev = mat.get("cantidad", 0.0)
                unidad_prev = mat.get("unidad_medida", "UN")
                actividad_prev = mat.get("actividad", "A elegir...")
                subactividad_prev = mat.get("subactividad", "")

                # --- 1) Preparamos opciones con placeholder ---
                opciones_codigo = [None] + list(catalogo_dict.keys())
                def format_codigo(x):
                    if x is None:
                        return "Seleccione…"
                    return f"{x} – {catalogo_dict[x]['descripcion']}"

                try:
                    index_codigo = opciones_codigo.index(codigo_prev)
                except ValueError:
                    index_codigo = 0

                # --- 2) Selectbox con placeholder ---
                codigo = col1.selectbox(
                    "Código",
                    options=opciones_codigo,
                    format_func=format_codigo,
                    key=f"mat_cod_{idx}",
                    index=index_codigo
                )
                # Capturar descripción
                if codigo is not None:
                    descripcion = catalogo_dict[codigo]["descripcion"]
                    col1.caption(descripcion)
                else:
                    descripcion = ""


                cantidad = col2.number_input(
                    "Cantidad",
                    min_value=0.0, step=1.0,
                    key=f"mat_cant_{idx}",
                    value=cantidad_prev
                )

                unidades = ["UN", "m", "GL"]
                # Si la descripción contiene "CABLE", preseleccionamos "m"
                if "CABLE" in descripcion.upper():
                    unidad_default = "m"
                else:
                    unidad_default = unidad_prev

                try:
                    index_unidad = unidades.index(unidad_default)
                except ValueError:
                    index_unidad = 0

                unidad = col3.selectbox(
                    "U.M.",
                    options=unidades,
                    index=index_unidad,
                    key=f"mat_um_{idx}"
                )

                # Botón para eliminar
                with col4:
                    # Dos saltos de línea para centrar verticalmente
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("❌", key=f"mat_del_{idx}", use_container_width=True):
                        materiales.pop(idx)
                        st.rerun()


                # Actividad principal
                actividad = st.selectbox(
                    "Actividad",
                    ["A elegir..."] + actividades,
                    index=(["A elegir..."] + actividades).index(actividad_prev) if actividad_prev in actividades else 0,
                    key=f"mat_act_{idx}"
                )
                if actividad in subactividades_dict:
                    subactividad = st.selectbox(
                        "Subactividad / Zona",
                        subactividades_dict[actividad],
                        index=subactividades_dict[actividad].index(subactividad_prev) if subactividad_prev in subactividades_dict[actividad] else 0,
                        key=f"mat_subact_{idx}"
                    )
                else:
                    subactividad = ""

                tipo = catalogo_dict.get(codigo, {}).get("tipo", "NO CLASIFICADO")
                nuevos_materiales.append({
                    "codigo": codigo,
                    "descripcion": catalogo_dict[codigo]["descripcion"] if codigo else "",
                    "cantidad": cantidad,
                    "unidad_medida": unidad,
                    "tipo": tipo,
                    "actividad": actividad,
                    "subactividad": subactividad
                })

            # Guardamos de nuevo en sesión
            st.session_state["materiales_registro"] = nuevos_materiales

            # 9. Vista previa
            st.markdown("### 📝 Vista previa de materiales cargados")
            if any(m["codigo"] for m in nuevos_materiales):
                df_mat = pd.DataFrame([m for m in nuevos_materiales if m["codigo"]])
                st.dataframe(df_mat, use_container_width=True)
            else:
                st.info("No hay materiales seleccionados.")

            # 10. Guardar materiales (pre-confirmación)
            if st.button("💾 Guardar materiales"):
                to_save = [m for m in nuevos_materiales if m["codigo"]]
                if not to_save:
                    st.warning("No se puede guardar: ningún material está seleccionado.")
                elif any(m["cantidad"] == 0 for m in to_save):
                    st.warning("No se puede guardar: todos los materiales deben tener cantidad mayor que cero.")
                elif any(m["actividad"] == "A elegir..." for m in to_save):
                    st.warning("No se puede guardar: todas las filas deben tener una actividad seleccionada.")
                else:
                    if existentes and not st.session_state.get("confirm_overwrite_mat", False):
                        st.warning("⚠️ Presioná de nuevo para sobrescribir los registros anteriores.")
                        st.session_state["confirm_overwrite_mat"] = True
                    else:
                        st.session_state["pendiente_confirmacion"] = True

            # 11. Confirmación final y operación de borrado + inserción
            # 11. Confirmación final y operación de borrado + inserción
            if st.session_state.get("pendiente_confirmacion", False):
                st.markdown("### 📤 Confirmar envío de materiales")
                
                # ——— Mostrar fecha y código de obra a guardar ———
                st.markdown(f"**Fecha:** {fecha_str}   |   **Código de obra:** {codigo_sel}")
                
                # Preparamos la lista de materiales a guardar
                to_save = [m for m in nuevos_materiales if m["codigo"]]
                st.dataframe(pd.DataFrame(to_save), use_container_width=True)

                if st.button("✅ Confirmar envío final"):
                    try:
                        if existentes:
                            delete_materiales_por_proyecto_fecha(proyecto_id, fecha_str)
                        for m in to_save:
                            sub = m['subactividad']
                            act = m['actividad']
                            act_sub = f"{sub}_{act}" if sub else act
                            add_material_usado({
                                "proyecto_id":    proyecto_id,
                                "fecha":          fecha_str,
                                "codigo":         m["codigo"],
                                "descripcion":    m["descripcion"],
                                "unidad_medida":  m["unidad_medida"],
                                "cantidad":       m["cantidad"],
                                "usuario":        username,
                                "tipo":           m["tipo"],
                                "actividad":      act_sub  
                            })
                        st.success("🎉 Materiales guardados correctamente.")

                        # 12. Preparar descarga
                        st.session_state["last_materiales_descarga"] = to_save
                        st.session_state["last_obra"]     = consulta_por_codigo(codigo_sel)[0][1]
                        st.session_state["last_cod_proj"] = codigo_sel
                        st.session_state["last_fecha"]    = fecha_reg

                        st.session_state["pendiente_confirmacion"] = False
                        st.session_state["confirm_overwrite_mat"]  = False
                        st.session_state["materiales_registro"]    = []

                    except Exception as e:
                        st.error(f"❌ Error al guardar: {e}")


                    if st.session_state.get("last_materiales_descarga"):
                        buffer = io.BytesIO()
                        generar_planilla_excel(
                            materiales=st.session_state["last_materiales_descarga"],
                            obra=st.session_state["last_obra"],
                            cod_proj=st.session_state["last_cod_proj"],
                            fecha=st.session_state["last_fecha"],
                            buffer=buffer
                        )
                        st.download_button(
                            label=f"📥 Descargar planilla de seguimiento del {fecha_reg}",
                            data=buffer.getvalue(),
                            file_name=f"planilla_{st.session_state['last_cod_proj']}_{fecha_reg}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )



        

        elif menu == "🔍 Consultar registros":
            st.title("🔍 CONSULTAR REGISTROS")
            
            # 1) Traer lista de códigos de proyecto
            codigos_unicos = consulta_proyectos()  # debe devolver List[str]
            
            if not codigos_unicos:
                st.warning("No hay proyectos registrados.")
                st.stop()
            
            # 2) Selector de código
            codigo_sel = st.selectbox("Seleccioná un código de proyecto:", sorted(codigos_unicos))
            
            # 3) Consultar registros para ese código
            registros = consulta_por_codigo(codigo_sel)  
            # registros: List[Tuple[11 elementos]]
            
            if not registros:
                st.warning(f"No se encontraron registros para '{codigo_sel}'.")
                st.stop()
            
            # 4) Mostrar resumen con solo Código, Cliente y Fecha
            resumen = [(r[0], r[1], r[2]) for r in registros]
            df_resumen = pd.DataFrame(resumen, columns=["Código", "Cliente", "Fecha"])
            st.dataframe(df_resumen, use_container_width=True)
            
            # 5) Selector de fecha para ver detalle
            fechas = df_resumen["Fecha"].astype(str).tolist()
            fecha_sel = st.selectbox("Ver detalle del día:", fechas)
            
            # 6) Desempaquetar y mostrar detalle completo
            idx = fechas.index(fecha_sel)
            (
                proyecto, cliente, fecha,
                clima_manana, clima_tarde,
                detalle, actividades, equipos,
                residente, encargado, usuario
            ) = registros[idx]

            st.markdown(f"### 📄 Detalle de **{codigo_sel}** — {fecha_sel}")

            # Metadatos en tres columnas
            m1, m2, m3 = st.columns(3)
            m1.markdown(f"**Cliente**\n\n{cliente}")
            m1.markdown(f"**Usuario**\n\n{usuario}")
            m2.markdown(f"**Clima Mañana**\n\n{clima_manana}")
            m2.markdown(f"**Clima Tarde**\n\n{clima_tarde}")
            m3.markdown(f"**Residente**\n\n{residente}")
            m3.markdown(f"**Encargado**\n\n{encargado}")

            # Detalle narrativo
            st.write(detalle)
            # Parseamos actividades si es string
            act_rows = []
            if isinstance(actividades, str):
                try:
                    actividades = json.loads(actividades)
                except json.JSONDecodeError:
                    st.error("❌ El campo 'actividades' no es un JSON válido")
                    actividades = []

            # Validamos que sea lista
            if not isinstance(actividades, list):
                st.error("⚠️ 'actividades' no es una lista válida")
                actividades = []

            # Recorremos actividades
            for i, act in enumerate(actividades):
                if isinstance(act, str):
                    try:
                        act = json.loads(act)
                    except json.JSONDecodeError:
                        st.warning(f"⛔ Actividad [{i}] no es JSON válido")
                        continue

                if not isinstance(act, dict) or "Roles" not in act:
                    st.warning(f"⚠️ Actividad [{i}] no tiene estructura esperada")
                    continue

                for act in actividades:
                    total_hh = sum(
                        vals.get("Personal", 0) * vals.get("HH", 0)
                        for vals in act["Roles"].values()
                    )
                    act_rows.append({
                        "Actividad":   act.get("Actividad", ""),
                        "Personal":  sum(vals.get("Personal", 0) for vals in act["Roles"].values()),
                        "HH Totales":total_hh
                    })


            df_act = pd.DataFrame(act_rows)
            df_act = df_act.drop_duplicates(subset="Actividad", keep="first").reset_index(drop=True)

            # 1) Si quieres eliminar donde PERSONAL y HH Totales sean 0:
            df_act = df_act[~((df_act["Personal"] == 0) & (df_act["HH Totales"] == 0))]

            # 2) (Opcional) Si además quieres quitar filas con valores NaN en ambas columnas:
            df_act = df_act.dropna(subset=["Personal", "HH Totales"], how="all")

            # Ahora sí ordenas
            df_act = df_act.sort_values(by="HH Totales", ascending=False).reset_index(drop=True)

            st.markdown(f"**🔧 HH del {fecha_sel}**")
            st.table(df_act)

            # --- EQUIPOS ---

            # Parseamos si viene como string
            if isinstance(equipos, str):
                try:
                    equipos = json.loads(equipos)
                    
                except json.JSONDecodeError:
                    st.error("❌ El campo 'equipos' no es un JSON válido")
                    equipos = []

            # Validamos que sea lista
            if not isinstance(equipos, list):
                st.error("⚠️ 'equipos' no es una lista válida")
                equipos = []

            df_eq = pd.DataFrame(equipos)
            st.markdown(f"**🚜 Equipos en obra del {fecha_sel}**")
            st.table(df_eq)
            

            # ——————————————————————————————————————————
            # 8) Gráfico Total HH por Actividad (Proyecto Completo)
            # ——————————————————————————————————————————
            import plotly.graph_objects as go
            total_data = {}
            for r in registros:
                acts = r[6]
                if isinstance(acts, str):
                    acts = json.loads(acts)
                for act in acts:
                    if isinstance(act, str):
                        act = json.loads(act)
                    if not isinstance(act, dict) or "Roles" not in act:
                        continue
                    nombre = act.get("Actividad", "")
                    hh = sum(v.get("Personal", 0) * v.get("HH", 0) for v in act["Roles"].values())
                    total_data[nombre] = total_data.get(nombre, 0) + hh

            # Actividades ordenadas
            actividades = list(total_data.keys())
            hh_totales = list(total_data.values())

            # ——————————————————————————————————————————
            # 9) Obtener metas desde Supabase por proyecto
            # ——————————————————————————————————————————
            proyecto_id = obtener_proyecto_id(codigo_sel)
            metas_dict = obtener_metas_hh(proyecto_id)

            # ——————————————————————————————————————————
            # 10) Formulario para modificar metas HH
            # ——————————————————————————————————————————
            st.subheader("🔧 Definir Meta HH por Actividad")

            act_sel = st.selectbox("Seleccioná la actividad para asignar Meta HH", actividades)
            meta_val = st.number_input(
                f"Meta HH – {act_sel}",
                min_value=0.0,
                step=0.5,
                value=metas_dict.get(act_sel, 0.0)
            )

            if st.button("Guardar Meta HH"):
                guardar_meta_hh(proyecto_id, act_sel, meta_val,username)
                st.success(f"Meta HH para “{act_sel}” guardada correctamente.")
                metas_dict[act_sel] = meta_val  # actualizamos en runtime

            import plotly.graph_objects as go

            # ——————————————————————————————————————
            # HH Totales vs Meta HH (con ducteado apilado, sin leyenda de sub-zonas)
            # ——————————————————————————————————————

            base_act    = "Ducteado Embutido/Endosado"
            sub_zonas   = ["MAMPOSTERIA", "LOSA", "SUBTERRANEO", "EXTERNO"]

            # 1) Acumular horas y metas
            zona_hh   = {z: 0.0 for z in sub_zonas}
            zona_meta = {z: 0.0 for z in sub_zonas}
            otros_acts, otros_hh, otros_metas = [], [], []

            for act, hh in zip(actividades, hh_totales):
                matched = False
                for z in sub_zonas:
                    if act == f"{z}-{base_act}":
                        zona_hh[z]   += hh
                        zona_meta[z]  = metas_dict.get(act, 0.0)
                        matched = True
                        break
                if not matched:
                    otros_acts.append(act)
                    otros_hh.append(hh)
                    otros_metas.append(metas_dict.get(act, 0.0))

            tot_hh_duct   = sum(zona_hh.values())
            tot_meta_duct = sum(zona_meta.values())

            # Contamos cuántas sub-zonas tuvieron HH > 0
            nonzero_zonas = sum(1 for hh in zona_hh.values() if hh > 0)

            # 2) Construir figura
            fig = go.Figure()

            # 2A) Cada sub-zona apilada, sin mostrar en leyenda
            for z in sub_zonas:
                fig.add_trace(go.Bar(
                    x=[base_act],
                    y=[zona_hh[z]],
                    name=z,
                    hovertemplate=f"{z}: %{{y:.1f}} HH<extra></extra>",
                    showlegend=False
                ))

            # 2B) El resto de actividades (barras individuales)
            fig.add_trace(go.Bar(
                x=otros_acts,
                y=otros_hh,
                marker=dict(color=otros_hh, colorscale='greens', showscale=False),
                text=otros_hh,
                textposition='auto',
                name="Otras Actividades",
                showlegend=False
            ))

            # 2C) Línea de Meta HH (única entrada en la leyenda)
            fig.add_trace(go.Scatter(
                x=[base_act] + otros_acts,
                y=[tot_meta_duct] + otros_metas,
                mode='lines+markers',
                name="Meta HH",
                line=dict(dash='dash', color='red'),
                marker=dict(size=8),
                hovertemplate="%{x}: %{y:.1f} HH<extra></extra>",
                showlegend=True
            ))

            # 3) Anotación del total de ducteado SOLO si hubo más de una sub-zona
            if nonzero_zonas > 1:
                fig.add_annotation(
                    x=base_act,
                    y=0,  # en la base de la pila
                    text=f"Total: {tot_hh_duct:.1f} HH",
                    showarrow=False,
                    yshift=10,    # sube el texto justo encima del eje
                    font=dict(size=12, color="black")
                )

            # 4) Ajustes de layout
            fig.update_layout(
                title="",
                barmode='stack',
                xaxis_title="Actividad",
                yaxis_title="Horas-Hombre",
                xaxis_tickangle=-45,
                bargap=0.2,
                legend=dict(
                    orientation='h',
                    yanchor='bottom',
                    y=1.02,
                    xanchor='right',
                    x=1
                ),
                margin=dict(t=80, b=150),
                template="plotly_white",
            )

            # 5) Mostrar en Streamlit dentro de un expander
            with st.expander("🏗️ HH Totales vs Meta HH por Actividad", expanded=True):
                st.plotly_chart(fig, use_container_width=True)

           # ——————————————————————————————————————————
            # 📉 Desvío de HH Totales vs Meta HH con Tolerancia (DENTRO DE EXPANDER)
            # ——————————————————————————————————————————
            with st.expander("📊 Desvío porcentual HH Totales vs Meta por Actividad", expanded=True):

                # 1) Input de tolerancia
                tolerancia = st.number_input(
                    "Tolerancia aceptable (%)", min_value=0.0, max_value=100.0, step=0.5, value=5.0
                )
                # 2) Calcular desvíos y colores según tolerancia
                desvio_data = []
                for act, hh_real in zip(actividades, hh_totales):
                    meta = metas_dict.get(act, 0.0)
                    if meta > 0 and hh_real > 0:
                        desvio = ((hh_real - meta) / meta) * 100
                        #abs_desvio = abs(desvio)

                        # Color según tolerancia
                        color = "#2ca02c" if desvio <= tolerancia else "#d62728"

                        desvio_data.append({
                            "Actividad": act,
                            "HH Totales": hh_real,
                            "Meta HH": meta,
                            "% Desvío": desvio,
                            "Color": color,
                            "Etiqueta": f"{desvio:.1f}%"
                        })

                # 3) Mostrar gráfico si hay datos válidos
                if not desvio_data:
                    st.info("No hay actividades con metas e información suficiente para mostrar desvío.")
                else:
                    df_desvio = pd.DataFrame(desvio_data)
                    df_desvio = df_desvio.sort_values(by="% Desvío", ascending=False)

                    fig_desvio = go.Figure()

                    fig_desvio.add_trace(go.Bar(
                        x=df_desvio["Actividad"],
                        y=df_desvio["% Desvío"],
                        marker_color=df_desvio["Color"],
                        text=df_desvio["Etiqueta"],
                        textposition="auto",  # usa inside/outside dinámico según espacio
                        insidetextanchor="middle",
                        insidetextfont=dict(color="white"),
                        hovertemplate=(
                            "<b>%{x}</b><br>" +
                            "Desvío: %{y:.1f}%<br>" +
                            "Meta HH: %{customdata[0]:,.0f}<br>" +
                            "HH Totales: %{customdata[1]:,.0f}<extra></extra>"
                        ),
                        customdata=df_desvio[["Meta HH", "HH Totales"]].values
                    ))

                    fig_desvio.update_layout(
                        title="",
                        xaxis_title="Actividad",
                        yaxis_title="% Desvío",
                        xaxis_tickangle=-45,
                        bargap=0.4,
                        template="plotly_white",
                        height=500,
                        margin=dict(t=80, b=140, l=60, r=40)
                    )

                    st.plotly_chart(fig_desvio, use_container_width=True)

            # ——————————————————————————————————————————
            # Gráfico de BARRAS apiladas: HH Diarias por Actividad
            # (se muestra solo si hay más de un día de datos)
            # ——————————————————————————————————————————
            # 1) Preparamos un dict fecha → {actividad: HH}
            data_by_date = {}
            for r in registros:
                fecha = pd.to_datetime(r[2]).date()
                acts = r[6]
                if isinstance(acts, str):
                    acts = json.loads(acts)
                for act in acts:
                    if isinstance(act, str):
                        act = json.loads(act)
                    if not isinstance(act, dict) or "Roles" not in act:
                        continue
                    nombre = act.get("Actividad", "")
                    hh = sum(vals.get("Personal", 0) * vals.get("HH", 0) for vals in act["Roles"].values())
                    data_by_date.setdefault(fecha, {}).setdefault(nombre, 0)
                    data_by_date[fecha][nombre] += hh

            # 2) Ordenamos fechas y actividades únicas
            fechas = sorted(data_by_date.keys())
            actividades = sorted({act for daily in data_by_date.values() for act in daily})

            # 3) Solo dibujamos si hay más de un día
            if len(fechas) > 1:
                fig = go.Figure()
                for actividad in actividades:
                    y = [data_by_date[fecha].get(actividad, 0) for fecha in fechas]
                    if all(v == 0 for v in y):
                        continue  # omitimos actividades sin HH en TODO el proyecto
                    fig.add_trace(go.Bar(
                        x=fechas,
                        y=y,
                        name=actividad,
                        hovertemplate="%{x}: %{y} HH<extra>" + actividad + "</extra>"
                    ))

                # 4) Layout para barras apiladas
                fig.update_layout(
                    barmode='stack',
                    title="",
                    xaxis=dict(
                        title="Fecha",
                        tickangle=-45,
                        type="category"
                    ),
                    yaxis=dict(title="HH Totales"),
                    legend=dict(
                        title="Actividad",
                        orientation="v",
                        yanchor="bottom",
                        y=1.02,
                        xanchor="right",
                        x=1
                    ),
                    template="plotly_white",
                    margin=dict(t=80, b=120, l=60, r=40)
                )

                # 5) Mostrar en Streamlit
                with st.expander("🕐Horas‑Hombre Diarias Apiladas por Actividad", expanded=True):
                    st.plotly_chart(fig, use_container_width=True)
            else:
                st.write("")

            excel_bytes = generar_excel_proyecto(codigo_sel)

            #único botón de descarga
            st.download_button(
                label="📥 Descargar REPORTE COMPLETO",
                data=excel_bytes,
                file_name=f"{codigo_sel}_resumen_{datetime.today().date()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        

        st.markdown(
                    "<div style='text-align: right; color: gray; font-size: 0.9em;'>Developed by Guillermo Ojeda Cueto</div>",
                    unsafe_allow_html=True
                )

        
    

    elif authentication_status is False:
        st.error("Usuario o contraseña incorrectos.")
    else:  # authentication_status is None
        st.warning("Por favor, ingresa tus credenciales.")
        if st.button("Registrarse"):
            st.session_state['click_en_registro'] = True
            st.rerun()


def confirm_msg():
    """Intenta crear el usuario. 
       Devuelve True si se agregó, False en caso contrario."""
    # 1) Verifico contraseñas
    if st.session_state['password'] != st.session_state['confirmar_password']:
        st.error("🔒 Las contraseñas no coinciden.")
        sleep(3)
        return False

    # 2) Verifico existencia previa
    if consulta_nombre(st.session_state['usuario']):
        st.warning("❗ El usuario ya existe.")
        sleep(3)
        return False

    # 3) Si todo OK, hago el registro
    hashed = stauth.Hasher([st.session_state['password']]).generate()[0]
    add_registro(
        st.session_state['nombre'].strip(),
        st.session_state['usuario'].strip(),
        hashed
    )
    st.success("✅ Usuario registrado correctamente.")
    sleep(3)
    return True


def usuario_form():
    with st.form(key='formulario', clear_on_submit=True):
        st.text_input("Nombre", key="nombre")
        st.text_input("Usuario", key="usuario")
        st.text_input("Contraseña",     type="password", key="password")
        st.text_input("Confirmar contraseña", type="password", key="confirmar_password")
        registrar = st.form_submit_button("Registrar")

    if registrar:
        # Validación de campos vacíos
        campos = [st.session_state['nombre'],
                  st.session_state['usuario'],
                  st.session_state['password'],
                  st.session_state['confirmar_password']]
        if not all(c.strip() for c in campos):
            st.error("❗ Por favor completa todos los campos antes de registrar.")
            return  # corto antes de llamar a confirm_msg()

        # Llamo a confirm_msg() y, si retorna True, puedo hacer otras cosas (o nada más)
        if confirm_msg():
            # aquí podrías redirigir, limpiar el estado, etc.
            pass

    # Botón para volver al login
    if st.button("Volver al Login"):
        st.session_state['click_en_registro'] = False
        st.rerun()

if __name__ == "__main__":
    main()


